# (IMPORTANTE):
# PARA O USO ADEQUADO PASSE ESSE SCRIPT PARA OUTRO LOCAL, PODE SER UMA NOVA ABA NO VSCODE E CLIQUE PARA EXECUTAR!
# 


from openpyxl import load_workbook
import pyautogui

# Definindo variaveis para poder acessar a planilha;
vendas_de_lanches = load_workbook('cadastro.xlsx')
pagina1 = vendas_de_lanches['Cadastro']

for linha in pagina1.iter_rows(min_row=2, values_only=True):
    # Nome
    pyautogui.click(552,441, duration=1.5)
    pyautogui.write(linha[0])
    # Produto
    pyautogui.click(554,462, duration=1.5)
    pyautogui.write(linha[1])
    # Quantidadeedro Santos
    pyautogui.click(555,483, duration=1.5)
    pyautogui.write(str(linha[2]))
    # Data
    pyautogui.click(555,506,duration=1.5)
    pyautogui.write(linha[3])

    # Salvar
    pyautogui.click(568,525, duration=1.5)
    pyautogui.click(285,307, duration=1.5)