# ESSE ARQUIVO É APENAS PARA TESTS, O ARQUIVO AONDE ENCONTRA O PROJETO É EM CONTABILIDADE!

from openpyxl import load_workbook
import pyautogui

# Definindo variaveis para poder acessar a planilha;
vendas_de_lanches = load_workbook('cadastro.xlsx')
pagina1 = vendas_de_lanches['Cadastro']

for linha in pagina1.iter_rows(min_row=2, values_only=True):
    # Nome
    pyautogui.click(606, 448, duration=1.5)
    # Produto
    print(linha[1])
    # Quantidade
    print(linha[2])
